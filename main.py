from aiogram import Bot, Dispatcher, executor, types
import datetime
import pandas as pd
import markup as nav
import Spisok as vav
from pycoingecko import CoinGeckoAPI
import openpyxl
bot = Bot(token=TOKEN)
dp = Dispatcher(bot)
cg = CoinGeckoAPI()
#из csv в xlsx
df = pd.read_csv('для моделирования.csv')
df.to_excel('для моделирования.xlsx', 'Sheet1',index=False)
df = pd.read_csv('материнская плата.csv')
df.to_excel('материнская плата.xlsx', 'Sheet1',index=False)
df = pd.read_csv('для игр.csv')
df.to_excel('для игр.xlsx', 'Sheet1',index=False)
df = pd.read_csv('для майнинга.csv')
df.to_excel('для майнинга.xlsx', 'Sheet1',index=False)
book = openpyxl.open('для моделирования.xlsx',read_only=True)
book1 = openpyxl.open('для игр.xlsx',read_only=True)
book2 = openpyxl.open('для майнинга.xlsx',read_only=True)
book3= openpyxl.open('материнская плата.xlsx',read_only=True)
sheet = book.worksheets[0]
sheet1 = book1.worksheets[0]
sheet2 = book2.worksheets[0]
sheet3 = book3.worksheets[0]
class Ore():
    j=0


@dp.message_handler(commands=['start'])
async def command_start(message: types.Message):
    text_f = open("Список пользователей.txt", 'a')
    mess_time = datetime.datetime.now()
    text_f.write(
        f"\n{message.text}-"f"{message.from_user.first_name} {message.from_user.last_name} ID:{message.from_user.id} DATA: {mess_time}")
    text_f.close()
    send_mess = f"Привет, {message.from_user.first_name} {message.from_user.last_name}!\nЯ бот от компании Hot apple pies\nВсегда готов помочь)))\nВыбери, пожалуйста, что тебе нужно."
    await message.answer_sticker(r'CAACAgIAAxkBAAEDG7dhbYuRCk6kEXB6Qw4CtBKAUFebmwACEQADOW9OJhAtX5wMXKDwIQQ')
    await bot.send_message(message.from_user.id, send_mess,
                           reply_markup=nav.mainMenu)


@dp.message_handler(lambda message: 'семья' in message.text)
async def ufo(message: types.Message):
    await message.answer_sticker(r'CAACAgIAAxkBAAEDYLlho6RK9fyfnDZeUAqVDBJTtw3DagACMw4AAqCJkEuIJLsbXsC1dCIE')


@dp.message_handler(lambda message: 'семьи' in message.text)
async def ufo(message: types.Message):
    await message.answer_sticker(r'CAACAgIAAxkBAAEDYLlho6RK9fyfnDZeUAqVDBJTtw3DagACMw4AAqCJkEuIJLsbXsC1dCIE')


@dp.message_handler(lambda message: 'Торетто' in message.text)
async def ufo(message: types.Message):
    await bot.send_message(message.from_user.id, 'Забирай всё!!!')
    await message.answer_sticker(r'CAACAgIAAxkBAAEDYLxho6X8iKFSJ8CEkgLxQcI65td8jwACggwAApBTgEsjNGMG89LWpCIE')
    keyboard1 = types.InlineKeyboardButton('Ссылка',
                                           url='https://www.citilink.ru/product/materinskaya-plata-asus-prime-b360m-k-lga-1151v2-intel-b360-matx-ret-1059814/')
    keyboard = types.InlineKeyboardMarkup()
    keyboard = keyboard.add(keyboard1)
    await bot.send_message(message.from_user.id,
                           'Заказ принят и оплачен \nЗдравствуйте, Павел! \nТовары уже готовятся к отправке — мы пришлем уведомление в день доставки. Отслеживать статус заказа в реальном времени можно в Личном кабинете.',
                           reply_markup=keyboard)


@dp.message_handler()
async def bot_message(message: types.Message):
    try:
        a='1'
        if message.text == 'Видеокарты':
            await bot.send_message(message.from_user.id, 'Супер!!!')
            await bot.send_message(message.from_user.id,
                                   'Теперь давай выберем: для каких задач тебе нужна видеокарта...',
                                   reply_markup=nav.videoMenu)
        elif message.text == 'Процессоры':
            await bot.send_message(message.from_user.id, 'Найс!!!')
            await bot.send_message(message.from_user.id,
                                   'Теперь давай выберем: для каких задач тебе нужен процессор...',
                                   reply_markup=nav.procMenu)
        elif message.text == 'Материнская плата':
            await bot.send_message(message.from_user.id, 'Крутяк!!!')
            await bot.send_message(message.from_user.id, 'Теперь скажи, пожалуйста: какой сокет у твоего процессора?',
                                   reply_markup=nav.matMenu)
        elif message.text == 'Я не знаю':
            keyboard1 = types.InlineKeyboardButton('Ссылка', url='https://www.youtube.com/watch?v=uHjAKz3gyPc')
            keyboard = types.InlineKeyboardMarkup()
            keyboard = keyboard.add(keyboard1)
            await bot.send_message(message.from_user.id,
                                   'Ничего страшного\nПосмотри видеоролик и возвращайся обратно\nЖду)))',
                                   reply_markup=keyboard)
        elif message.text == 'Для игр':
            await bot.send_message(message.from_user.id, 'Игры - это всегда круто)))')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.videogameMenu)
        elif message.text == 'Для майнинга':
            await bot.send_message(message.from_user.id, 'Крутой)))')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.videomineMenu)
        elif message.text == 'Для моделирования':
            await bot.send_message(message.from_user.id, 'Вау!!!')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.videomodMenu)
        elif message.text == 'LGA 1700':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.price1700Menu)
        elif message.text == 'LGA 1200':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: какой чипсет у твоего процессора',
                                   reply_markup=nav.e1200Menu)
        elif message.text == 'SocketAM4':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: какой чипсет у твоего процессора',
                                   reply_markup=nav.eAM4Menu)
        elif message.text == 'LGA 1151v2':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: какой чипсет у твоего процессора',
                                   reply_markup=nav.price1151Menu)
        elif message.text == 'Intel B365':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eB365Menu)
        elif message.text == 'Intel H310':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eH310Menu)
        elif message.text == 'Intel H310C':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eH310CMenu)
        elif message.text == 'Intel B460':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eB460Menu)
        elif message.text == 'Intel B560':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eB560Menu)
        elif message.text == 'Intel H410':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eH410Menu)
        elif message.text == 'Intel H470':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eH470Menu)
        elif message.text == 'Intel H510':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eH510Menu)
        elif message.text == 'Intel H570':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eH570Menu)
        elif message.text == 'Intel Z590':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eZ590Menu)
        elif message.text == 'AMD A320':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eAM320Menu)
        elif message.text == 'AMD A520':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eAM520Menu)
        elif message.text == 'AMD B450':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eAM450Menu)
        elif message.text == 'AMD B550':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eAM550Menu)
        elif message.text == 'AMD X570':
            await bot.send_message(message.from_user.id, 'Окей')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.eAM570Menu)
        elif message.text == 'спешиал фор game💥':
            await bot.send_message(message.from_user.id, 'Замечательно)))')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.pricegameprocMenu)
        elif message.text == 'Для майнинга💥':
            await bot.send_message(message.from_user.id, 'Хорошо)))')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.pricemineprocMenu)
        elif message.text == 'Для моделирования💥':
            await bot.send_message(message.from_user.id, 'Хорошо)))')
            await bot.send_message(message.from_user.id, 'Теперь выбери: в каком промежутке находится твой бюджет?',
                                   reply_markup=nav.pricemodprocMenu)
        # формула для моделирования
        elif message.text == '8000-15000':
            await bot.send_message(message.from_user.id, 'Ищем...')
            a = False
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) >= 8000 and int(sheet[f"B{i}"].value) <= 15000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '15000-20000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 15000 and int(sheet[f"B{i}"].value) < 20000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '20000-25000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 20000 and int(sheet[f"B{i}"].value) < 25000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '25000-30000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 25000 and int(sheet[f"B{i}"].value) < 30000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '30000-35000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 30000 and int(sheet[f"B{i}"].value) < 35000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '35000-40000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 35000 and int(sheet[f"B{i}"].value) < 40000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '40000-45000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 40000 and int(sheet[f"B{i}"].value) < 45000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '45000-55000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 45000 and int(sheet[f"B{i}"].value) < 55000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '55000-70000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 55000 and int(sheet[f"B{i}"].value) < 70000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '70000-90000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 70000 and int(sheet[f"B{i}"].value) < 90000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '90000-150000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 90000 and int(sheet[f"B{i}"].value) < 150000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '150000-210000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 150000 and int(sheet[f"B{i}"].value) < 210000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '210000-300000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 210000 and int(sheet[f"B{i}"].value) < 300000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 300000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                if b[0] == 'X' or b[0] == 'E':
                    if int(sheet[f"B{i}"].value) > 300000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'до 30000👾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                f1 = f.replace(', ', '\n')
                if b[0] == 'Q':
                    if int(sheet[f"B{i}"].value) < 30000 and int(sheet[f"B{i}"].value):
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '30000-40000👾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                f1 = f.replace(', ', '\n')
                if b[0] == 'Q':
                    if int(sheet[f"B{i}"].value) > 30000 and int(sheet[f"B{i}"].value) < 40000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '40000-60000👾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                f1 = f.replace(', ', '\n')
                if b[0] == 'Q':
                    if int(sheet[f"B{i}"].value) > 40000 and int(sheet[f"B{i}"].value) < 60000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '60000-80000👾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                f1 = f.replace(', ', '\n')
                if b[0] == 'Q':
                    if int(sheet[f"B{i}"].value) > 60000 and int(sheet[f"B{i}"].value) < 80000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '80000-100000👾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                f1 = f.replace(', ', '\n')
                if b[0] == 'Q':
                    if int(sheet[f"B{i}"].value) > 80000 and int(sheet[f"B{i}"].value) < 100000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '100000-150000👾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                f1 = f.replace(', ', '\n')
                if b[0] == 'Q':
                    if int(sheet[f"B{i}"].value) > 100000 and int(sheet[f"B{i}"].value) < 150000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '150000-250000👾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                f1 = f.replace(', ', '\n')
                if b[0] == 'Q':
                    if int(sheet[f"B{i}"].value) > 150000 and int(sheet[f"B{i}"].value) < 250000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 250000👾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet.max_row + 1):
                b = str(sheet[f"A{i}"].value)
                h = str(sheet[f"D{i}"].value)
                v = str(sheet[f"C{i}"].value)
                f = str(sheet[f"E{i}"].value)
                f1 = f.replace(', ', '\n')
                if b[0] == 'Q':
                    if int(sheet[f"B{i}"].value) > 250000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:],f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        # для игр
        elif message.text == '10000-13000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 10000 and int(sheet1[f"B{i}"].value) < 13000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '13000-20000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 13000 and int(sheet1[f"B{i}"].value) < 20000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '20000-25000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 20000 and int(sheet1[f"B{i}"].value) < 25000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '25000-30000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 25000 and int(sheet1[f"B{i}"].value) < 30000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '30000-35000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 30000 and int(sheet1[f"B{i}"].value) < 35000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '35000-40000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 35000 and int(sheet1[f"B{i}"].value) < 40000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '40000-45000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                print(v, h)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 40000 and int(sheet1[f"B{i}"].value) < 45000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '45000-50000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 45000 and int(sheet1[f"B{i}"].value) < 50000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '50000-60000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 50000 and int(sheet1[f"B{i}"].value) < 60000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '60000-70000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 60000 and int(sheet1[f"B{i}"].value) < 70000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 70000💢':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'I':
                    if int(sheet1[f"B{i}"].value) > 70000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'до 30000❗':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] != 'I':
                    if int(sheet1[f"B{i}"].value) > 18000 and int(sheet1[f"B{i}"].value) < 30000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '30000-40000❗':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] != 'I' and b[0] != 'D':
                    if int(sheet1[f"B{i}"].value) > 30000 and int(sheet1[f"B{i}"].value) < 40000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '40000-60000❗':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] != 'I':
                    if int(sheet1[f"B{i}"].value) > 40000 and int(sheet1[f"B{i}"].value) < 60000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '60000-80000❗':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] != 'I':
                    if int(sheet1[f"B{i}"].value) > 60000 and int(sheet1[f"B{i}"].value) < 80000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '80000-100000❗':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] != 'I':
                    if int(sheet1[f"B{i}"].value) > 80000 and int(sheet1[f"B{i}"].value) < 100000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '100000-150000❗':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] != 'I':
                    if int(sheet1[f"B{i}"].value) > 100000 and int(sheet1[f"B{i}"].value) < 150000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '150000-250000❗':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] != 'I':
                    if int(sheet1[f"B{i}"].value) > 150000 and int(sheet1[f"B{i}"].value) < 250000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 250000❗':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet1.max_row + 1):
                b = str(sheet1[f"A{i}"].value)
                h = str(sheet1[f"D{i}"].value)
                v = str(sheet1[f"C{i}"].value)
                f = str(sheet1[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] != 'I':
                    if int(sheet1[f"B{i}"].value) > 250000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet1[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        # для майнинга
        elif message.text == '12000-15000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'A' and b[1] == 'M' and b[2] == 'D':
                    if int(sheet2[f"B{i}"].value) > 12000 and int(sheet2[f"B{i}"].value) < 15000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '15000-20000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'A' and b[1] == 'M' and b[2] == 'D':
                    if int(sheet2[f"B{i}"].value) > 15000 and int(sheet2[f"B{i}"].value) < 20000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '20000-25000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'A' and b[1] == 'M' and b[2] == 'D':
                    if int(sheet2[f"B{i}"].value) > 20000 and int(sheet2[f"B{i}"].value) < 25000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '25000-30000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'A' and b[1] == 'M' and b[2] == 'D':
                    if int(sheet2[f"B{i}"].value) > 25000 and int(sheet2[f"B{i}"].value) < 30000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '30000-35000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'A' and b[1] == 'M' and b[2] == 'D':
                    if int(sheet2[f"B{i}"].value) > 30000 and int(sheet2[f"B{i}"].value) < 35000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '35000-45000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'A' and b[1] == 'M' and b[2] == 'D':
                    if int(sheet2[f"B{i}"].value) > 35000 and int(sheet2[f"B{i}"].value) < 45000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '45000-55000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'A' and b[1] == 'M' and b[2] == 'D':
                    if int(sheet2[f"B{i}"].value) > 45000 and int(sheet2[f"B{i}"].value) < 55000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '55000-70000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'A' and b[1] == 'M' and b[2] == 'D':
                    if int(sheet2[f"B{i}"].value) > 55000 and int(sheet2[f"B{i}"].value) < 70000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == 'свыше 70000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[0] == 'A' and b[1] == 'M' and b[2] == 'D':
                    if int(sheet2[f"B{i}"].value) > 70000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == 'до 30000⚠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[:3] != 'AMD':
                    if int(sheet2[f"B{i}"].value) > 19000 and int(sheet2[f"B{i}"].value) < 30000 and b[len(b) - 4] != 'L' and b[len(b) - 4] != 'L':
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '30000-40000⚠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[:3] != 'AMD' and b[0] != 'D':
                    if int(sheet2[f"B{i}"].value) > 30000 and int(sheet2[f"B{i}"].value) < 40000 and b[len(b) - 4] != 'L':
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '40000-60000⚠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[:3] != 'AMD':
                    if int(sheet2[f"B{i}"].value) > 40000 and int(sheet2[f"B{i}"].value) < 60000 and b[len(b) - 4] != 'L':
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '60000-80000⚠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[:3] != 'AMD':
                    if int(sheet2[f"B{i}"].value) > 60000 and int(sheet2[f"B{i}"].value) < 80000 and b[len(b) - 4] != 'L':
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '80000-100000⚠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[:3] != 'AMD':
                    if int(sheet2[f"B{i}"].value) > 80000 and int(sheet2[f"B{i}"].value) < 100000 and b[len(b) - 4] != 'L':
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '100000-150000⚠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[:3] != 'AMD':
                    if int(sheet2[f"B{i}"].value) > 100000 and int(sheet2[f"B{i}"].value) < 150000 and b[len(b) - 4] != 'L':
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == '150000-250000⚠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[:3] != 'AMD':
                    if int(sheet2[f"B{i}"].value) > 150000 and int(sheet2[f"B{i}"].value) < 250000 and (b[len(b) - 4] != 'L' and b[len(b) - 7] != 'L'):
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        elif message.text == 'свыше 250000⚠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                b = str(sheet2[f"A{i}"].value)
                h = str(sheet2[f"D{i}"].value)
                v = str(sheet2[f"C{i}"].value)
                f = str(sheet2[f"E{i}"].value)
                f1 = f.replace('|', '\n')
                if b[:3] != 'AMD':
                    if int(sheet2[f"B{i}"].value) > 250000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(f'{sheet2[f"A{i}"].value}',
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.minemainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.minemainMenu)
        # материнские платы
        elif message.text == '6000-7000◀':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB460':
                    if int(sheet3[f"B{i}"].value) > 6000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '7000-9000◀':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB460':
                    if int(sheet3[f"B{i}"].value) > 7000 and int(sheet3[f"B{i}"].value) < 9000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '9000-11000◀':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB460':
                    if int(sheet3[f"B{i}"].value) > 9000 and int(sheet3[f"B{i}"].value) < 11000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 11000◀':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB460':
                    if int(sheet3[f"B{i}"].value) > 11000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '6000-7000◽':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB560':
                    if int(sheet3[f"B{i}"].value) > 6000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '7000-9000◽':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB560':
                    if int(sheet3[f"B{i}"].value) > 7000 and int(sheet3[f"B{i}"].value) < 9000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '9000-10000◽':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB560':
                    if int(sheet3[f"B{i}"].value) > 9000 and int(sheet3[f"B{i}"].value) < 10000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '10000-12000◽':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB560':
                    if int(sheet3[f"B{i}"].value) > 10000 and int(sheet3[f"B{i}"].value) < 12000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '12000-15000◽':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB560':
                    if int(sheet3[f"B{i}"].value) > 12000 and int(sheet3[f"B{i}"].value) < 15000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '15000-17000◽':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB560':
                    if int(sheet3[f"B{i}"].value) > 15000 and int(sheet3[f"B{i}"].value) < 17000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 17000◽':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB560':
                    if int(sheet3[f"B{i}"].value) > 17000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '5000-6000◾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH410':
                    if int(sheet3[f"B{i}"].value) > 5000 and int(sheet3[f"B{i}"].value) < 6000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '6000-7000◾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH410':
                    if int(sheet3[f"B{i}"].value) > 6000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 7000◾':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH410':
                    if int(sheet3[f"B{i}"].value) > 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '5000-6000💠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH470':
                    if int(sheet3[f"B{i}"].value) > 5000 and int(sheet3[f"B{i}"].value) < 6000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '6000-7000💠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet2.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH470':
                    if int(sheet3[f"B{i}"].value) > 6000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 7000💠':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH470':
                    if int(sheet3[f"B{i}"].value) > 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '5000-6000🔳':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH510':
                    if int(sheet3[f"B{i}"].value) > 5000 and int(sheet3[f"B{i}"].value) < 6000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '6000-7000🔳':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH510':
                    if int(sheet3[f"B{i}"].value) > 6000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 7000🔳':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH510':
                    if int(sheet3[f"B{i}"].value) > 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '9000-11000🔲':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH570':
                    if int(sheet3[f"B{i}"].value) > 9000 and int(sheet3[f"B{i}"].value) < 11000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '11000-12000🔲':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH570':
                    if int(sheet3[f"B{i}"].value) > 11000 and int(sheet3[f"B{i}"].value) < 12000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '12000-14000🔲':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH570':
                    if int(sheet3[f"B{i}"].value) > 12000 and int(sheet3[f"B{i}"].value) < 14000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 14000🔲':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH570':
                    if int(sheet3[f"B{i}"].value) > 14000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '8000-12000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ590':
                    if int(sheet3[f"B{i}"].value) > 8000 and int(sheet3[f"B{i}"].value) < 12000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '12000-14000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ590':
                    if int(sheet3[f"B{i}"].value) > 12000 and int(sheet3[f"B{i}"].value) < 14000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '14000-15000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ590':
                    if int(sheet3[f"B{i}"].value) > 14000 and int(sheet3[f"B{i}"].value) < 15000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '15000-17000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ590':
                    if int(sheet3[f"B{i}"].value) > 15000 and int(sheet3[f"B{i}"].value) < 17000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '17000-18000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ590':
                    if int(sheet3[f"B{i}"].value) > 17000 and int(sheet3[f"B{i}"].value) < 18000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '18000-22000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ590':
                    if int(sheet3[f"B{i}"].value) > 18000 and int(sheet3[f"B{i}"].value) < 22000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '22000-27000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ590':
                    if int(sheet3[f"B{i}"].value) > 22000 and int(sheet3[f"B{i}"].value) < 27000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '27000-36000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ590':
                    if int(sheet3[f"B{i}"].value) > 27000 and int(sheet3[f"B{i}"].value) < 36000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 36000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ590':
                    if int(sheet3[f"B{i}"].value) > 36000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '3500-4000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDA320':
                    if int(sheet3[f"B{i}"].value) > 3500 and int(sheet3[f"B{i}"].value) < 4000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '4000-5000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDA320':
                    if int(sheet3[f"B{i}"].value) > 4000 and int(sheet3[f"B{i}"].value) < 5000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '5000-6000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDA320':
                    if int(sheet3[f"B{i}"].value) > 5000 and int(sheet3[f"B{i}"].value) < 6000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 6000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDA320':
                    if int(sheet3[f"B{i}"].value) > 6000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '4000-5000🔋':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDA520':
                    if int(sheet3[f"B{i}"].value) > 4000 and int(sheet3[f"B{i}"].value) < 5000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '5000-6000🔋':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDA520':
                    if int(sheet3[f"B{i}"].value) > 5000 and int(sheet3[f"B{i}"].value) < 6000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '6000-7000🔋':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDA520':
                    if int(sheet3[f"B{i}"].value) > 6000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 7000🔋':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDA520':
                    if int(sheet3[f"B{i}"].value) > 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '4000-5000💯':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB450':
                    if int(sheet3[f"B{i}"].value) > 4000 and int(sheet3[f"B{i}"].value) < 5000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '5000-6000💯':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB450':
                    if int(sheet3[f"B{i}"].value) > 5000 and int(sheet3[f"B{i}"].value) < 6000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '6000-7000💯':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB450':
                    if int(sheet3[f"B{i}"].value) > 6000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '7000-8000💯':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB450':
                    if int(sheet3[f"B{i}"].value) > 7000 and int(sheet3[f"B{i}"].value) < 8000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '8000-9000💯':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB450':
                    if int(sheet3[f"B{i}"].value) > 8000 and int(sheet3[f"B{i}"].value) < 9000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 9000💯':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB450':
                    if int(sheet3[f"B{i}"].value) > 9000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '▶6000-7000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB550':
                    if int(sheet3[f"B{i}"].value) > 6000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '▶7000-9000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB550':
                    if int(sheet3[f"B{i}"].value) > 7000 and int(sheet3[f"B{i}"].value) < 9000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '▶9000-11000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB550':
                    if int(sheet3[f"B{i}"].value) > 9000 and int(sheet3[f"B{i}"].value) < 11000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '▶11000-14000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB550':
                    if int(sheet3[f"B{i}"].value) > 11000 and int(sheet3[f"B{i}"].value) < 14000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '▶14000-20000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB550':
                    if int(sheet3[f"B{i}"].value) > 14000 and int(sheet3[f"B{i}"].value) < 20000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '▶свыше 20000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB550':
                    if int(sheet3[f"B{i}"].value) > 20000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '12000-14000🔋':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDX570':
                    if int(sheet3[f"B{i}"].value) > 12000 and int(sheet3[f"B{i}"].value) < 14000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '14000-17000🔋':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDX570':
                    if int(sheet3[f"B{i}"].value) > 14000 and int(sheet3[f"B{i}"].value) < 17000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '17000-21000🔋':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDX570':
                    if int(sheet3[f"B{i}"].value) > 17000 and int(sheet3[f"B{i}"].value) < 21000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '21000-30000🔋':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDX570':
                    if int(sheet3[f"B{i}"].value) > 21000 and int(sheet3[f"B{i}"].value) < 30000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 30000🔋':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDX570':
                    if int(sheet3[f"B{i}"].value) > 30000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '4000-5000🔲':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB365':
                    if int(sheet3[f"B{i}"].value) > 4000 and int(sheet3[f"B{i}"].value) < 5000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '5000-6000🔲':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB365':
                    if int(sheet3[f"B{i}"].value) > 5000 and int(sheet3[f"B{i}"].value) < 6000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '6000-7000🔲':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB365':
                    if int(sheet3[f"B{i}"].value) > 6000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 7000🔲':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB365':
                    if int(sheet3[f"B{i}"].value) > 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '3500-4000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH310':
                    if int(sheet3[f"B{i}"].value) > 3500 and int(sheet3[f"B{i}"].value) < 4000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '4000-5000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH310':
                    if int(sheet3[f"B{i}"].value) > 4000 and int(sheet3[f"B{i}"].value) < 5000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '5000-7000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH310':
                    if int(sheet3[f"B{i}"].value) > 5000 and int(sheet3[f"B{i}"].value) < 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 7000✅':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH310':
                    if int(sheet3[f"B{i}"].value) > 7000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '▶3500-4000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH310C':
                    if int(sheet3[f"B{i}"].value) > 3500 and int(sheet3[f"B{i}"].value) < 4000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '▶4000-6000':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH310C':
                    if int(sheet3[f"B{i}"].value) > 4000 and int(sheet3[f"B{i}"].value) < 6000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '17000-20000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ690':
                    if int(sheet3[f"B{i}"].value) > 17000 and int(sheet3[f"B{i}"].value) < 20000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '20000-25000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ690':
                    if int(sheet3[f"B{i}"].value) > 20000 and int(sheet3[f"B{i}"].value) < 25000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '25000-30000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ690':
                    if int(sheet3[f"B{i}"].value) > 25000 and int(sheet3[f"B{i}"].value) < 30000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '30000-35000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ690':
                    if int(sheet3[f"B{i}"].value) > 30000 and int(sheet3[f"B{i}"].value) < 35000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '35000-40000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ690':
                    if int(sheet3[f"B{i}"].value) > 35000 and int(sheet3[f"B{i}"].value) < 40000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == '40000-50000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ690':
                    if int(sheet3[f"B{i}"].value) > 40000 and int(sheet3[f"B{i}"].value) < 50000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'свыше 50000📱':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ690':
                    if int(sheet3[f"B{i}"].value) > 50000:
                        a = True
                        keyboard1 = types.InlineKeyboardButton(g[17:],
                                                               url=v[1:])
                        keyboard = types.InlineKeyboardMarkup()
                        keyboard = keyboard.add(keyboard1)
                        await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                             reply_markup=keyboard)
                        await bot.send_message(message.from_user.id,
                                               'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'Intel Z490':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ490':
                    a = True
                    keyboard1 = types.InlineKeyboardButton(g[17:],
                                                           url=v[1:])
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard = keyboard.add(keyboard1)
                    await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                         reply_markup=keyboard)
                    await bot.send_message(message.from_user.id,
                                           'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'AMD B350':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDB350':
                    a = True
                    keyboard1 = types.InlineKeyboardButton(g[17:],
                                                           url=v[1:])
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard = keyboard.add(keyboard1)
                    await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                         reply_markup=keyboard)
                    await bot.send_message(message.from_user.id,
                                           'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'AMD X370':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDX370':
                    a = True
                    keyboard1 = types.InlineKeyboardButton(g[17:],
                                                           url=v[1:])
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard = keyboard.add(keyboard1)
                    await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                         reply_markup=keyboard)
                    await bot.send_message(message.from_user.id,
                                           'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'AMD X470':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'AMDX470':
                    a = True
                    keyboard1 = types.InlineKeyboardButton(g[17:],
                                                           url=v[1:])
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard = keyboard.add(keyboard1)
                    await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                         reply_markup=keyboard)
                    await bot.send_message(message.from_user.id,
                                           'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'Intel B360':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelB360':
                    a = True
                    keyboard1 = types.InlineKeyboardButton(g[17:],
                                                           url=v[1:])
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard = keyboard.add(keyboard1)
                    await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                         reply_markup=keyboard)
                    await bot.send_message(message.from_user.id,
                                           'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'Intel H370':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelH370':
                    a = True
                    keyboard1 = types.InlineKeyboardButton(g[17:],
                                                           url=v[1:])
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard = keyboard.add(keyboard1)
                    await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                         reply_markup=keyboard)
                    await bot.send_message(message.from_user.id,
                                           'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'Intel Z390':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelZ390':
                    a = True
                    keyboard1 = types.InlineKeyboardButton(g[17:],
                                                           url=v[1:])
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard = keyboard.add(keyboard1)
                    await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                         reply_markup=keyboard)
                    await bot.send_message(message.from_user.id,
                                           'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'LGA 2066':
            a = False
            await bot.send_message(message.from_user.id, 'Ищем...')
            for i in range(2, sheet3.max_row + 1):
                h = str(sheet3[f"D{i}"].value)
                v = str(sheet3[f"C{i}"].value)
                f = str(sheet3[f"E{i}"].value)
                g = str(sheet3[f"A{i}"].value)
                f1 = f.replace('|', '\n')
                if str(sheet3[f"F{i}"].value) == 'IntelX299':
                    a = True
                    keyboard1 = types.InlineKeyboardButton(g[17:],
                                                           url=v[1:])
                    keyboard = types.InlineKeyboardMarkup()
                    keyboard = keyboard.add(keyboard1)
                    await bot.send_photo(message.from_user.id, h[1:], f1[1:],
                                         reply_markup=keyboard)
                    await bot.send_message(message.from_user.id,
                                           'Идёт поиск...\nПожалуйста, дождитесь окончания операции')
            if a == True:
                await bot.send_message(message.from_user.id,
                                       'Поиск завершен)\nСпасибо, что воспользовались нашим сервисом😊😊😊',
                                       reply_markup=nav.smainMenu)
            else:
                await bot.send_message(message.from_user.id,
                                       'По данным параметрам нет ничего в наличии(((\nМожет быть посмотрите что-то другое?',
                                       reply_markup=nav.smainMenu)
        elif message.text == 'В главное меню':
            await message.answer_sticker(r'CAACAgIAAxkBAAEDYHRho4jqQfyxJ0ONRF7-QzpSkPvDpgACHgADwDZPE6FgWy2rAAHeBCIE')
            send_mess = "Выбери, пожалуйста, что тебе нужно)"
            await bot.send_message(message.from_user.id, send_mess,
                                   reply_markup=nav.mainMenu)
        elif message.text == 'restart':
            await message.answer_sticker(r'CAACAgIAAxkBAAEDYHJho4iNCaH9IObNWsWxcRlmm6d2ZQACHAADwDZPE8GCGtMs_g7hIgQ')
            await bot.send_message(message.from_user.id, 'Хэй!!!')
            await bot.send_message(message.from_user.id, 'Чем могу еще помочь?)))', reply_markup=nav.mainMenu)
        elif message.text == '/leavefeedback':
            Ore.j = 1
            text_f = open("bool отзыв.txt", 'w')
            text_f.write(
                f"\n{message.text}-"f"{message.from_user.first_name} {message.from_user.last_name} ID:{message.from_user.id}")
            text_f.close()
            await bot.send_message(message.from_user.id,
                                   'Теперь можешь написать сам отзыв)',
                                   reply_markup=nav.feedbackMenu)
        elif Ore.j==1:
            if message.text:
                text_f = open("Отзывы.txt", 'a')
                text_f.write(
                    f"\n{message.text}-"f"{message.from_user.first_name} {message.from_user.last_name} ID:{message.from_user.id}")
                text_f.close()
                await bot.send_message(message.from_user.id,
                                       'Спасибо тебе за отзыв!!!\nТвое мнение важно для нашей команды!!!\nВсегда будем рады помочь тебе)))',
                                       reply_markup=nav.vmainMenu)
                Ore.j = 0
        elif message.text == 'Плохо' or message.text == 'Хороший бот' or message.text == 'Всё супер':
            text_f = open("Отзывы.txt", 'a')
            text_f.write(
                f"\n{message.text}-"f"{message.from_user.first_name} {message.from_user.last_name} ID:{message.from_user.id}")
            text_f.close()
            await bot.send_message(message.from_user.id,
                                   'Спасибо тебе за отзыв!!!\nТвое мнение важно для нашей команды!!!\nВсегда будем рады помочь тебе)))',
                                   reply_markup=nav.vmainMenu)
        elif message.text == '/help':
            await bot.send_message(message.from_user.id, vav.spisok, reply_markup=nav.vmainMenu)
        elif message.text == '/mine':
            await bot.send_message(message.from_user.id, 'Выберете криптовалюту, курс которой ты хочешь узнать?',
                                   reply_markup=nav.crupto_list)
        elif message.text[0] == '/' and message.text[1] == '/':
            result = cg.get_price(ids=message.text[2:], vs_currencies='usd')
            await bot.send_message(message.from_user.id,
                                   f"Криптовалюта: {message.text[2:]}\nСтоимость на данный момент: {result[message.text[2:]]['usd']}$")
        else:
            await bot.send_message(message.from_user.id,
                                   'Я не знаю таких команд...\nПожалуйста, воспользуйся панельками🤗\nЧтобы посмотреть список моих возможностей нажми на /help')
    except Exception as err:
        text_f = open("Список возникших ошибок.txt", 'a')
        mess_time1 = datetime.datetime.now()
        text_f.write(
            f"\n{err} - {message.text}-"f"{message.from_user.first_name} {message.from_user.last_name} ID:{message.from_user.id} - DATA: {mess_time1}")
        text_f.close()
        await bot.send_message(message.from_user.id,
                               'Во время загрузки что-то пошло не так(((\nПопробуйте сделать запрос позже, либо можете выбрать что-то другое',
                               reply_markup=nav.vmainMenu)


@dp.callback_query_handler(text_contains="cc_")
async def crupto(call: types.CallbackQuery):
    await bot.delete_message(call.from_user.id, call.message.message_id)
    callback_data = call.data
    currency = str(callback_data[3:])
    result = cg.get_price(ids=currency, vs_currencies='usd')
    await bot.send_message(call.from_user.id,
                           f"Криптовалюта: {currency}\nСтоимость на данный момент: {result[currency]['usd']}$", reply_markup=nav.crupto_list)


if __name__ =='__main__':
    executor.start_polling(dp, skip_updates = True)



